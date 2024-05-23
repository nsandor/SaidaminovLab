import matplotlib.pyplot as plt
import pandas as pd
from glob import glob
import os
import numpy as np
from scipy.optimize import curve_fit
from sklearn.metrics import r2_score
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

def list_directories(path):
    path = Path(path)
    # Get all elements in a folder and store only directories in the list
    directories = [child for child in path.iterdir() if child.is_dir()]
    directories.sort()
    return directories

def single_exponential_decay(t,A1,tau1, c):
    return A1 * np.exp((-1)*t/tau1) + c

def bi_exponential_decay(t,A1,A2,tau1,tau2, c):
    return A1 * np.exp((-1)*t/tau1) + A2 * np.exp((-1)*t/tau2) + c

def tri_exponential_decay(t,A1,A2,A3,tau1,tau2,tau3, c):
    return A1 * np.exp((-1)*t/tau1) + A2 * np.exp((-1)*t/tau2) + A3 * np.exp((-1)*t/tau3) + c

def process_data(X,Y,normalize=True):
    # find on-time
    on_index = Y.index(max(Y))
    offset_time = X[on_index]
    X = [x-offset_time for x in X]
    if normalize:
        # normalize
        Y = [(y-min(Y))/max(Y) for y in Y]
    return X, Y, on_index

def take_a_look(X,Y,label,log_plot=True):
    plt.scatter(X,Y,c='white',ec='blue',label=label)
    plt.xlim(0,None)
    if log_plot:
        plt.yscale('log')
    plt.xlabel('Time (ns)')
    plt.ylabel('Counts')
    plt.legend(frameon=False)
    plt.show()
    return

def check_data(path):
    data = pd.read_csv(path,delimiter=',',skiprows=10,header=None)
    label = os.path.basename(path)[:-4]
    X = data[0].to_list() # time (ns)
    Y = data[1].to_list() # count (cts)
    X, Y, _ = process_data(X, Y, normalize=True)
    take_a_look(X,Y,label,log_plot=True)

def show_fit_result(popt,r2,r2_log,results,label,X,Y,X_for_fit,Y_fit,figsave,fig_dir):
    print('Fit: a1=%5.3f, a2=%5.3f, a3=%5.3f, t1=%5.3f, t2=%5.3f, t3=%5.3f, c=%5.3f' % tuple(popt))
    print(f'R2: {r2:.4f}, R2_log: {r2_log:.4f}')

    lifetime = max(results[3:6])

    fig = plt.figure()
    plt.scatter(X, Y, c='white', ec='#00939280')
    plt.plot(X_for_fit, Y_fit, c='#009392',label=f'{label}: {lifetime:.1f} ns')
    plt.ylabel('Counts (a.u.)')
    plt.xlabel('Time (ns)')
    plt.ylim(1e-4, 1)
    plt.xlim(0, None)
    plt.yscale('log')
    plt.legend(frameon=False)

    if figsave:
        # Make directory if you don't have it
        if not os.path.exists(fig_dir):
            os.mkdir(fig_dir)
        fig_save_path = f'{fig_dir}/{label}.png'
        plt.savefig(fig_save_path)

    plt.show()

    return fig

def TRPL_fit(X, Y, label, fig_dir, figsave=False):
    # Make On time to be 0 ns and Normalize Y
    X, Y, on_index = process_data(X, Y, normalize=True)

    # Decide cut time based on the figure
    # take_a_look(X, Y,label)
    # cut_time = int(input('Cut time (ns)?'))

    # Cut time for fitting
    power = -3
    threshold = 10 ** power
    while True:
        try:
            cut_index = next(y[0] for y in enumerate(Y[on_index:]) if y[1] < threshold) + on_index
            break
        except:
            power += 0.2
            threshold = 10 ** power
    cut_time = X[cut_index]
    print(f'{cut_time:.1f} ns')

    # Data used for fitting
    X_for_fit = X[on_index:cut_index]
    Y_for_fit = Y[on_index:cut_index]
    # Prepare log for fitting
    log_Y_for_fit = np.log(Y_for_fit)

    try:
        # start from tri exponential
        def log_fit(t, A1, A2, A3, tau1, tau2, tau3, c):
            return np.log(tri_exponential_decay(t, A1, A2, A3, tau1, tau2, tau3, c))
        
        # Initial guess for the parameters
        initial_guess = [0.7, 0.2, 0.1, 5, 50, 100, 0]
        # Bounds for the parameters (all parameters are positive)
        bounds = (0, np.inf)  # All parameters must be greater than 0

        # Fit the tri-exponential function to the data with bounds
        popt, _ = curve_fit(log_fit, X_for_fit, log_Y_for_fit, p0=initial_guess, bounds=bounds)

        # Calculate R-squared value
        Y_fit = [np.exp(log_fit(x, popt[0],popt[1],popt[2],popt[3],popt[4],popt[5], popt[6])) for x in X_for_fit]
        r2 = r2_score(Y_for_fit, Y_fit)
        r2_log = r2_score(log_Y_for_fit, np.log(Y_fit))

        results = list(popt)

        if min(results[0:2]) > 0.01:
            show_fit_result(popt,r2,r2_log,results,label,X,Y,X_for_fit,Y_fit,figsave,fig_dir)

        else:
            # Fit with bi-exponential
            def log_fit(t, A1, A2, tau1, tau2,c):
                return np.log(bi_exponential_decay(t, A1, A2, tau1, tau2,c))
            
            # Initial guess for the parameters
            initial_guess = [0.7, 0.2, 5, 50, 0]
            # Bounds for the parameters (all parameters are positive)
            bounds = (0, np.inf)  # All parameters must be greater than 0

            # Fit the tri-exponential function to the data with bounds
            popt, _ = curve_fit(log_fit, X_for_fit, log_Y_for_fit, p0=initial_guess, bounds=bounds)

            # Calculate R-squared value
            Y_fit = [np.exp(log_fit(x, popt[0],popt[1],popt[2],popt[3],popt[4])) for x in X_for_fit]
            r2 = r2_score(Y_for_fit, Y_fit)
            r2_log = r2_score(log_Y_for_fit, np.log(Y_fit))

            results = list(popt)
            results = [results[0],results[1],0,results[2],results[3],0,results[4]]

            if min(results[0:1]) > 0.01:
                show_fit_result(popt,r2,r2_log,results,label,X,Y,X_for_fit,Y_fit,figsave,fig_dir)

            else:
                # Fit with single-exponential
                def log_fit(t, A1, tau1,c):
                    return np.log(single_exponential_decay(t, A1, tau1,c))
                
                # Initial guess for the parameters
                initial_guess = [0.7, 0.2, 0]
                # Bounds for the parameters (all parameters are positive)
                bounds = (0, np.inf)  # All parameters must be greater than 0

                # Fit the tri-exponential function to the data with bounds
                popt, _ = curve_fit(log_fit, X_for_fit, log_Y_for_fit, p0=initial_guess, bounds=bounds)

                # Calculate R-squared value
                Y_fit = [np.exp(log_fit(x, popt[0],popt[1],popt[2])) for x in X_for_fit]
                r2 = r2_score(Y_for_fit, Y_fit)
                r2_log = r2_score(log_Y_for_fit, np.log(Y_fit))

                results = list(popt)
                results = [results[0],0,0,results[1],0,0,0]

                show_fit_result(popt,r2,r2_log,results,label,X,Y,X_for_fit,Y_fit,figsave,fig_dir)

    except Exception as e:
        print(f'Fitting failed due to {e}')
        results = [0,0,0,0,0,0,0]
        r2, r2_log = 0, 0

    # round values
    cut_time = round(cut_time,1)
    results1 = [round(n,4) for n in results[:3]] # A1, A2, A3
    results2 = [round(n,2) for n in results[3:6]] # t1, t2, t3
    results3 = [round(results[-1],4)] # c
    results = results1 + results2 + results3
    r2 = round(r2,4)
    r2_log = round(r2_log,4)
    return cut_time, results, r2, r2_log

# Excel design
def Design_excel(excel_path, font = "Meiryo UI", fontsize = 10, head_bkg_color = '000000', head_let_color = 'FFFFFF'):
    """
    Load excel file:
    1. Set the font for all cells to "Meiryo UI"
    2. Set the style for the first row (Fill black, make bold and white)
    3. Freeze the top row
    4. Adjust column widths
    5. Align center
    """

    # Load the Excel file
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    # Set the font for all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name=font, size=fontsize)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    # Set the style for the first row
    for cell in ws['1:1']:
        cell.fill = PatternFill(start_color=head_bkg_color, end_color=head_bkg_color, fill_type="solid")
        cell.font = Font(color=head_let_color, bold=True, name=font, size=fontsize)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    # Freeze the top row
    ws.freeze_panes = 'A2'

    # Adjust column widths based on content
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Adjust the width factor as necessary
        ws.column_dimensions[column].width = adjusted_width
    # Save the changes
    wb.save(excel_path)

