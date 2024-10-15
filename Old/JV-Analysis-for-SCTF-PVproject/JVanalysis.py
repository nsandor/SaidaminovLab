'''
JV-analysis for Vishal's project
'''

from glob import glob
import os
from pathlib import Path
from datetime import datetime, timedelta
import pandas as pd
import matplotlib.pyplot as plt
import csv
import numpy as np
import tkinter as tk
from tkinter import filedialog
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def list_directories(path):
    path = Path(path)
    # Get all elements in a folder and store only directories in the list
    directories = [child for child in path.iterdir() if child.is_dir()]
    directories.sort()
    return directories

# Remove unnecessary folder
def clean_folder_list(folder_list, skip_indices):
    """Remove some directories from a folder list"""
    skip_indices = [n-1 for n in skip_indices]
    # Sort index in descending order
    sorted_indices = sorted(skip_indices, reverse=True)
    for index in sorted_indices:
        folder_list.pop(index)
    return folder_list

def start_analysis(folder_path):
    """Get folder list"""
    # Show the contents in the folder list
    folder_list = list_directories(folder_path)
    for i, folder in enumerate(folder_list):
        print(f'{i+1:02}: {os.path.basename(folder)}')
    # Remove unnecessary folder
    skip_str = input('Remove folders? Yes (0) or No (1)')
    if skip_str == '0':
        skip_indices_str = input('Choose folder IDs. Ex: 7-8, or 1-2-5, or 8')
        skip_indices = [int(x) for x in skip_indices_str.split('-')]
        skip_indices.sort()
        folder_list = clean_folder_list(folder_list, skip_indices)
        print('\nFolder list updated')
        for i, folder in enumerate(folder_list):
            print(f'{i+1:02}: {os.path.basename(folder)}')
    return folder_list  

def calculate_date_difference(date1, date2, date_format="%y%m%d"):
    """Calculate the difference in days between two dates given in YYMMDD format."""
    d1 = datetime.strptime(str(date1), date_format)
    d2 = datetime.strptime(str(date2), date_format)
    return abs((d2 - d1).days)

def add_days_to_date(date_str, days):
    # convert date_str to datetime object
    date = datetime.strptime(str(date_str), '%y%m%d')
    # add the days
    new_date = date + timedelta(days=days)
    # convert the date to str
    new_date = int(new_date.strftime('%y%m%d'))
    return new_date

# Define hex id for each cell
def get_cell_hex_id(dat):    # expecting dat file path in dat
    """
    Make hex id based on the data
    
    ex: CD240309 FD240325 Batch1 Sample01 Cell02 -> 24030924032510102 -> '555FFF82950896'
    to decode, use int(N,16)
    the decoded number is [6-digit][6-digit][1-digit][2-digit][2-digit] = [crystallization_date][fabrication_date][batch_number][sample_id][cell_id]
    unless you use this code until June 2072, the digits for hex id is always 14.
    """

    # Have a different format to get directory info
    dat_path_info  = Path(dat) # have a different format for later

    crystallization_date = dat_path_info.parents[3].name[1:7] # ex: take 240329 from 'C240329'
    fabrication_date = dat_path_info.parents[2].name.split('_')[0][1:7] # ex: take 240329 from 'F240329_B1'
    batch_number = dat_path_info.parents[2].name[-1:] # last letter in the folder name
    sample_id = int(dat_path_info.parents[0].name.split('_')[0])
    try:
        cell_id = int(os.path.basename(dat).split('-')[0])
    except:
        cell_id = int(os.path.basename(dat).split('_')[0])
    cell_hex_id = hex(int(f'{crystallization_date}{fabrication_date}{batch_number}{sample_id:02}{cell_id:02}'))[2:].upper()

    return cell_hex_id

# Decode cell_hex_id
def decode_cell_hex_id(cell_hex_id):
    # Convert hex to decimal
    cell_info = str(int(cell_hex_id,16))
    
    crystallization_date = cell_info[:6]
    fabrication_date = cell_info[6:12]
    batch_number = cell_info[12:13]
    sample_id = cell_info[13:15]
    cell_id = cell_info[15:17]

    return [int(crystallization_date), int(fabrication_date), int(batch_number), int(sample_id), int(cell_id)]

# Excel design
def Design_excel(excel_path, font = "Meiryo UI", fontsize = 10, head_bkg_color = '000000', head_let_color = 'FFFFFF'):
    """
    Load excel file:
    1. Set the font for all cells to "Meiryo UI"
    2. Set the style for the first row (Fill black, make bold and white)
    3. Freeze the top row
    4. Adjust column widths
    5. Align center
    """

    # Load the Excel file
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    # Set the font for all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name=font, size=fontsize)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    # Set the style for the first row
    for cell in ws['1:1']:
        cell.fill = PatternFill(start_color=head_bkg_color, end_color=head_bkg_color, fill_type="solid")
        cell.font = Font(color=head_let_color, bold=True, name=font, size=fontsize)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    # Freeze the top row
    ws.freeze_panes = 'A2'

    # Adjust column widths based on content
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Adjust the width factor as necessary
        ws.column_dimensions[column].width = adjusted_width
    # Save the changes
    wb.save(excel_path)

def fill_excel_bkg(excel_path, indication_number):
    # fill bkg color based on the sample name
    bkg_colors = ["F2F2F2", "D9D9D9"] # light gray, gray
        
    # Load the Excel file
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    # Fill the background
    color_change = 0
    for i, row in enumerate(ws.iter_rows()):
        if i == 0:
            pass
        elif i == 1:
            indicator = row[indication_number].value # define sample number
            for cell in row:
                cell.fill = PatternFill(start_color=bkg_colors[0], end_color=bkg_colors[0], fill_type="solid") # use light gray
        elif i > 1:
            if indicator != row[indication_number].value:
                indicator = row[indication_number].value # update
                color_change += 1
                for cell in row:
                    cell.fill = PatternFill(start_color=bkg_colors[color_change % 2], end_color=bkg_colors[color_change % 2], fill_type="solid")
            else:
                for cell in row:
                    cell.fill = PatternFill(start_color=bkg_colors[color_change % 2], end_color=bkg_colors[color_change % 2], fill_type="solid")

    wb.save(excel_path)


def get_cell_information(dat, df_cell_info, mode):
    error_check = []
    cols = df_cell_info.columns.to_list()
    # Get all cell info
    try:
        # Define hex_id
        cell_hex_id = get_cell_hex_id(dat)
        if cell_hex_id in df_cell_info['ID'].values: # if the dataframe has the cell info, skip
            pass
        else:
            # get some info based on the directory name
            crystallization_date, fabrication_date, batch_number, sample_id, cell_id = decode_cell_hex_id(cell_hex_id)
            # get other info
            sample_info_parts = Path(dat).parents[0].name.split('_')
            sample_info = '_'.join(sample_info_parts[1:])
            area_dat = float(pd.read_csv(dat,delimiter='\t')[:1]['device area'][0])

            # Define new row for the dataframe
            new_cell_info = [[
                        cell_hex_id,
                        crystallization_date,
                        fabrication_date,
                        batch_number,
                        sample_id,
                        sample_info,
                        cell_id,
                        area_dat,
                        None, # corrected area
                        '' # Note
                    ]]

            new_row = pd.DataFrame(data=new_cell_info, columns=cols)

            if mode == 'Create New':
                if len(df_cell_info) == 0:
                    df_cell_info = new_row
                else:
                    df_cell_info = pd.concat([df_cell_info, new_row], ignore_index=True, axis=0)

            elif mode == 'Update':
                # Merge the previous data and new row based on ID
                merged_df = df_cell_info.set_index('ID').combine_first(new_row.set_index('ID'))
                # For 'Area_new' and 'Selection', we use the existing data
                merged_df['Area_new'] = df_cell_info.set_index('ID')['Area_new'].combine_first(merged_df['Area_new'])
                merged_df['Note'] = df_cell_info.set_index('ID')['Note'].combine_first(merged_df['Note'])
                merged_df = merged_df.reset_index()[df_cell_info.columns]
                # Update df_cell_info
                df_cell_info = merged_df

    except Exception as e:
        error_check.append(f'error in {os.path.basename(dat)}, "{e}" ')

    return error_check, df_cell_info
            
# Grab Cell Information and Update/Create the summary as an Excel file
def update_cell_information(folder_list, save_path):
    mode = 'Update' # 'Update' or 'Create New'
    # if the excel file already exists
    try:
        # Read the existing excel file
        df_cell_info = pd.read_excel(save_path)
        cols = df_cell_info.columns.to_list()
    except:
        # Empty dataframe for data
        cols = ['ID','C-date','F-date','Batch','Sample','Info','Cell','Area_dat','Area_new','Note']
        df_cell_info = pd.DataFrame(index=[], columns=cols)
        mode = 'Create New'

    for folder in folder_list:
        try:
            # grab all dat files in the directory
            dat_list = glob(f'{folder}/*.dat')
            dat_list.sort()
            
            # get cell information and update dataframe
            for dat in dat_list:
                error_check, df_cell_info = get_cell_information(dat, df_cell_info, mode)

            # Message
            if len(error_check) == 0:
                pass
            elif len(error_check) == len(dat_list):
                print(' PASS, no data was saved\n')
            else:
                for error in error_check:
                    print(f'CAUTION, {os.path.basename(folder)}: {error}')

        except Exception as e:
            print(f'Skipped "{os.path.basename(folder)}" due to the error "{e}"')

    # Save the data
    df_cell_info = df_cell_info.sort_values(by='ID').reset_index(drop=True)
    with pd.ExcelWriter(save_path) as writer:
        df_cell_info.to_excel(writer, sheet_name='Cell-Info', index=False)

    # Design
    Design_excel(save_path)

    # fill bkg color based on the sample name
    fill_excel_bkg(save_path,indication_number=4)

    print('COMPLETE')

    return

def get_jv_id(dat):
    dat_folder = os.path.dirname(dat)
    dat_list = glob(f'{dat_folder}/*.dat')
    hex_id = get_cell_hex_id(dat)
    # get some info based on the directory name
    fabrication_date = decode_cell_hex_id(hex_id)[1] # int
    measurement_date = Path(dat).parents[1].name[1:7] # ex: take 240329 from 'M240329'
    # calculate how many days have passed
    measurement_days = calculate_date_difference(fabrication_date,measurement_date)
    # Count which trial the dat file is
    dat_list.sort()
    trial = 1
    for file in dat_list:
        file_hex_id = get_cell_hex_id(file) # get hex key to check the file is same as the objective file
        if file_hex_id == hex_id and file == dat: # if it is the same cell and same measurement data
            break
        elif file_hex_id == hex_id and file != dat: # if it is the same cell but different measurement data
            trial += 1
        else:
            pass
    # Define jv_id
    jv_id = f'{hex_id}-{measurement_days}-{trial:02}'
    return jv_id

def make_new_filename_dat2csv(dat):
    # Define hex_id
    cell_hex_id = get_cell_hex_id(dat)
    # get some info based on the directory name
    crystallization_date, fabrication_date, batch_number, sample_id, cell_id = decode_cell_hex_id(cell_hex_id)
    measurement_info = Path(dat).name[2:-4]
    measurement_date = Path(dat).parents[1].name[1:7] # ex: take 240329 from 'M240329'
    sample_info = Path(dat).parents[0].name

    # calculate how many days have passed
    fabrication_days = calculate_date_difference(crystallization_date,fabrication_date)
    measurement_days = calculate_date_difference(fabrication_date,measurement_date)
    
    # scan_direction
    df = pd.read_csv(dat,delimiter='\t',skiprows=2,header=None) # read dataframe
    if df.at[1, 0] > df.at[2, 0]: # if the first voltage is higher than the second voltage, direction is 'reverse (R)'
        scan_direction = 'R'
    else:
        scan_direction = 'F'
        
    # define new filename
    filename = f'{crystallization_date}_FD{fabrication_days}-B{batch_number}_{sample_info}-{cell_id:02}_MD{measurement_days}-{scan_direction}_{measurement_info}.csv'

    return filename


def save_JV_as_csv(V, I, J, file_save_path):
    # save to csv file
    with open(file_save_path, mode='w', newline='') as file:
        writer = csv.writer(file)
        # header
        writer.writerow(["Voltage (V)", "Current (mA)", "Current density (mA/cm2)"])
        # data
        for v, i, j in zip(V, I, J):
            writer.writerow([v, i, j])

def save_dat_as_csv(dat, save_folder_path, performance_csv_path):
    """
    read dat file
    1. convert it to csv file and rename it
    2. make jv-id (id for each measurement) and update performance summary as csv
    """

    # Error check
    error_check = []

    # get the data
    jv_id = get_jv_id(dat)

    df = pd.read_csv(dat,delimiter='\t',skiprows=2,header=None) # read dataframe
    df_performance = pd.read_csv(dat,delimiter='\t')[:1]
    area = float(df_performance['device area'][0]) # get area from original data
    J = df[1].to_list()
    J = [n*(-1) for n in J] # current density (mA/cm2)
    I = [n*area for n in J] # current (mA)
    V = df[0].to_list() # voltage (V)

    # Add jv-id
    df_performance['ID'] = jv_id
    # Add csv name
    df_performance['CSV name'] = make_new_filename_dat2csv(dat)
    # Change order
    new_order = ['ID', 'CSV name'] + [col for col in df_performance.columns if col not in ['ID', 'CSV name']]
    df_performance = df_performance[new_order]

    # Merge
    try:
        df_performance_csv = pd.read_csv(performance_csv_path)
        # Merge the previous data and new row based on ID
        merged_df = df_performance_csv.set_index('ID').combine_first(df_performance.set_index('ID'))
        merged_df = merged_df.reset_index()[df_performance_csv.columns]
        merged_df.to_csv(performance_csv_path, index=False)
    except:
        df_performance.to_csv(performance_csv_path, index=False)
    
    try:
        # define new filename
        filename = make_new_filename_dat2csv(dat)
        file_save_path = f'{save_folder_path}/{filename}'

        # save to csv file
        save_JV_as_csv(V, I, J, file_save_path)

    except Exception as e:
        error_check.append(f'Error in {os.path.basename(dat)}: "{e}"')
    return error_check

def save_all_dat_as_csv(folder_list, save_folder_dir, performance_csv_path):
    # Make directory if you don't have it
    if not os.path.exists(save_folder_dir):
        os.mkdir(save_folder_dir)

    for folder in folder_list:
        try:
            # grab all dat files in the directory
            dat_list = glob(f'{folder}/*.dat')
            dat_list.sort()
            # save all JV data
            for dat in dat_list:
                error_check = save_dat_as_csv(dat, save_folder_dir, performance_csv_path)
            if len(error_check) == 0:
                pass
            elif len(error_check) == len(dat_list):
                print(' PASS, no data was saved\n')
            else:
                print(f'CAUTION: {len(error_check)} of {len(dat_list)} dat files in {os.path.basename(folder)} were not saved. Check below.')
                for error in error_check:
                    print(error)

        except Exception as e:
            print(f' PASS, due to the error "{e}"\n')
    
    print('\nCOMPLETE')

def insert_hyperlink(excel_path,jv_folder_dir,performance_csv_path):
    df_reference =  pd.read_csv(performance_csv_path)
    
    # Open an existing Excel file
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Locate 'ID' column (assuming ID column is always present here)
    id_column_index = None
    for col in ws[1]:  # check the header
        if col.value == 'ID':
            id_column_index = col.column  # Get an index for ID
    
    # Determine the column letter for 'ID' column using its index
    id_column_letter = get_column_letter(id_column_index)

    # Add hyperlinks to the 'ID' column
    for row in range(2, ws.max_row + 1):  # Assuming the first row is the header
        cell = ws[f"{id_column_letter}{row}"]
        # Define hyperlink
        jv_id = cell.value # Get jv id ex 555F73CFCB9895-35-02
        csv_name = df_reference.loc[df_reference['ID']==jv_id, 'CSV name'].values[0]
        link = f"{jv_folder_dir}/{csv_name}"
        # Change cell settings
        cell.hyperlink = link
        # cell.style = "Hyperlink"

    # Save the changes and overwrite the existing file
    wb.save(excel_path)


def update_cell_performances(params):
    '''
    Update cell perfomance summary based on cell-information data frame
    '''
    jv_folder_dir = params[0]
    performance_csv_path = params[1]
    cell_info_path = params[2]
    performance_summary_path = params[3]

    # Read measurement data summary
    try:
        df_reference =  pd.read_csv(performance_csv_path)
        measurement_ids = df_reference['ID'].values
    except:
        print('Could not read the csv. Please check the path.')
        return

    # Read cell-information summary
    try:
        df_cell_info = pd.read_excel(cell_info_path)
    except:
        print('Could not read Cell information. Please check the file.')
        return

    # if the excel file already exists
    try:
        # Read the existinf excel file
        df_performances = pd.read_excel(performance_summary_path)
        cols = df_performances.columns.to_list()
        mode = 'Update'
    except:
        # Empty dataframe for data
        cols = ['ID','Selected','C-date','F-date','F-days','Batch','Sample','Info','Cell','Area (cm2)','Area correction','Trial','Filename','M-date','M-days','Scan',
                'Jsc (mA/cm2)','Voc (V)','FF (%)','Pmax (mW/cm2)','Vmpp (V)','Rseries(ohm)','Rshunt(ohm)','Note']
        df_performances = pd.DataFrame(index=[], columns=cols)
        mode = 'Create New'

    for jv_id in measurement_ids:
        # Get reference dataframe
        df_measurement = df_reference[df_reference['ID'] == jv_id]
        # Get cell id to get other cell information
        cell_hex_id = jv_id.split('-')[0]
        crystallization_date, fabrication_date, batch_number, sample_id, cell_id = decode_cell_hex_id(cell_hex_id)
        fabrication_days = calculate_date_difference(crystallization_date,fabrication_date)
        sample_info = df_cell_info[df_cell_info['ID'] == cell_hex_id]['Info'].values[0]

        # Get area based on cell-information
        area_dat = df_cell_info[df_cell_info['ID'] == cell_hex_id]['Area_dat'].values[0]
        area_correction = 0 # will be updated to 1 if you correct the area
        # Check if the dataframe has a corrected area
        nan_check = np.isnan(df_cell_info.loc[df_cell_info['ID'] == cell_hex_id, 'Area_new'].values[0])
        if nan_check: # if the value is nan
            area = area_dat
        else:
            area = df_cell_info[df_cell_info['ID'] == cell_hex_id]['Area_new'].values[0]
            area_correction = 1
        
        # trial number
        trial = int(jv_id.split('-')[2])
        # measurement filename
        filename = df_measurement['File Name'].values[0][2:-4]

        # Get measurement information based on jv-id
        measurement_days = int(jv_id.split('-')[1])
        measurement_date = add_days_to_date(date_str=fabrication_date,days=measurement_days)

        # get j-v data
        csv_name = df_measurement['CSV name'].values[0]
        csv_path = f'{jv_folder_dir}/{csv_name}'
        df_jv = pd.read_csv(csv_path)

        V = df_jv['Voltage (V)'].values
        I = df_jv['Current (mA)'].values
        J = df_jv['Current density (mA/cm2)'].values

        # update j-v data
        area_original_dat = df_measurement['device area'].values[0]
        if area_original_dat != area:
            J = [n/area for n in I]
            save_JV_as_csv(V, I, J, csv_path)

        # scan direction
        # check the scan direction based on the voltage data
        if V[0] > V[-1]:
            scan_direction = 'R'
        else:
            scan_direction = 'F'
        
        # get performance info
        Jsc = df_measurement['Jsc (mA/cm2)'].values[0] * area_original_dat * (-1) / area # mA/cm2: conver it to I then use correct area
        Voc = df_measurement['Voc (V)'].values[0] # V
        FF = df_measurement['Fill Factor'].values[0] * 100 # %
        Pmax = df_measurement['Pmax (mW/cm2)'].values[0] * area_original_dat * (-1) / area
        Vmpp = df_measurement['Vmpp'].values[0] # V
        Rseries = df_measurement['Rseries (Ohms)'].values[0] # Ohms
        Rshunt = df_measurement['Rshunt (Ohms)'].values[0] # Ohms

        # Define new row for the dataframe
        new_performance = [[
                    jv_id,
                    0, # selected
                    crystallization_date,
                    fabrication_date,
                    fabrication_days,
                    batch_number,
                    sample_id,
                    sample_info,
                    cell_id,
                    area,
                    area_correction,
                    trial,
                    filename,
                    measurement_date,
                    measurement_days,
                    scan_direction,
                    Jsc,
                    Voc,
                    FF,
                    Pmax,
                    Vmpp,
                    Rseries,
                    Rshunt,
                    '' # Note
                ]]

        new_row = pd.DataFrame(data=new_performance, columns=cols)

        # Merge
        if mode == 'Create New':
            if len(df_performances) == 0:
                df_performances = new_row
            else:
                df_performances = pd.concat([df_performances, new_row], ignore_index=True, axis=0)

        elif mode == 'Update':
            # Merge the previous data and new row based on ID
            merged_df = df_performances.set_index('ID').combine_first(new_row.set_index('ID'))
            # For 'Area_new' and 'Selection', we use the existing data
            merged_df['Selected'] = df_performances.set_index('ID')['Selected'].combine_first(merged_df['Selected'])
            merged_df['Note'] = df_performances.set_index('ID')['Note'].combine_first(merged_df['Note'])
            merged_df = merged_df.reset_index()[df_performances.columns]
            # Update df_performances
            df_performances = merged_df
            
    # Save the data
    with pd.ExcelWriter(performance_summary_path) as writer:
        df_performances.to_excel(writer, sheet_name='Performances', index=False)

    # Insert hyperlink for csv files
    insert_hyperlink(performance_summary_path,jv_folder_dir,performance_csv_path)
    # Design Excel
    Design_excel(performance_summary_path)
    # fill bkg color based on the cell id
    fill_excel_bkg(performance_summary_path,indication_number=8)

    print('COMPLETE')

    return


# Select the best JV curves
def select_best_curve(params):
    folder_list = params[0]
    jv_folder_dir = params[1]
    cell_info_path = params[2]
    performance_summary_path = params[3]
    performance_csv_path = params[4]
    df_performances = pd.read_excel(performance_summary_path) # Read summary excel
    df_cell_info = pd.read_excel(cell_info_path)

    cell_hex_ids = [] # Recode which cell was analyzed
    select_ids = []

    for folder in folder_list:
        # check which folder you are operating
        print(f'{os.path.basename(folder)}\n')

        # grab all dat files in the directory
        dat_list = glob(f'{folder}/*.dat')
        dat_list.sort()

        # get know how many cells you have
        data_id_list = []
        for dat in dat_list:
            try:
                # get info from file name
                dat_name = os.path.basename(dat)
                try:
                    cell_id = float(dat_name.split('-')[0])
                except:
                    cell_id = float(dat_name.split('_')[0])
                cell_hex_id = get_cell_hex_id(dat)
                cell_hex_ids.append(cell_hex_id)
                data_id_list.append(cell_id)
            except:
                # remove exceptional data
                print(f'CAUTION: skipped {dat_name}')
                dat_list.remove(dat)
                continue
        check_list = list(set(data_id_list))
        check_list.sort()

        # plot all data to select the best JV for each cell
        for check_id in check_list:
            print(f'\nCell {int(check_id):02}')
            fig = plt.figure(figsize=(6,4))

            for dat in dat_list:
                # get info from file name
                dat_name = os.path.basename(dat)
                try:
                    cell_id = float(dat_name.split('-')[0])
                except:
                    cell_id = float(dat_name.split('_')[0])
                jv_id = get_jv_id(dat)
                # "Selected" will be 0 for all cells that were checked (initialize)
                df_performances.loc[df_performances['ID'] == jv_id, 'Selected'] = 0
                
                if cell_id == check_id:
                    # Get some info
                    measurement_id = jv_id[:-2] # ex '555F73CFCB9895-35-'
                    measurement_info = dat_name[2:-4]
                    csv_name = make_new_filename_dat2csv(dat) # get csv name
                    csv_path = f'{jv_folder_dir}/{csv_name}'
                    trial = int(jv_id.split('-')[2])

                    # check performance
                    df_performance = df_performances[df_performances['ID'] == jv_id]
                    Jsc = float(df_performance['Jsc (mA/cm2)'].values[0])
                    Voc = float(df_performance['Voc (V)'].values[0])
                    FF = float(df_performance['FF (%)'].values[0])
                    Pmax = float(df_performance['Pmax (mW/cm2)'].values[0])
                    area_original = float(df_performance['Area (cm2)'].values[0])

                    # get J-V data
                    df = pd.read_csv(csv_path)
                    V = df['Voltage (V)'].values
                    I = df['Current (mA)'].values

                    # Use area from cell_info
                    area_from_cell_info = df_cell_info[df_cell_info['ID']==jv_id.split('-')[0]]['Area_dat'].values[0]
                    if df_performance['Area correction'].values[0] == 1:
                        area_from_cell_info = df_cell_info[df_cell_info['ID']==jv_id.split('-')[0]]['Area_new'].values[0]
                    # Correct data based on the area in cell info
                    Jsc = Jsc * area_original / area_from_cell_info
                    Pmax = Pmax * area_original / area_from_cell_info
                    J = [n/area_from_cell_info for n in I]

                    # plot
                    plt.plot(V,J,label=trial)

                    print(f'#{trial:02}  Jsc: {Jsc:.2f} mA/cm2, Voc: {Voc:.3f} V, FF: {FF:.1f}%, Pmax: {Pmax:.3f} mW/cm2, {measurement_info}')

            # figure design
            plt.xlabel('Voltage (V)')
            plt.ylabel('Current density (mA/cm2)')
            plt.xlim(0,None)
            plt.ylim(0,None)
            plt.legend(frameon=False)
            plt.show()

            while True:
                X = input(f'Select a number for the best JV data of Cell {int(check_id):02}')
                try:
                    print(f'#{X} was selected for Cell {int(check_id):02}')
                    # get selected jv-id for the selected data
                    select_id = f'{measurement_id}{int(X):02}'
                    if select_id in df_performances['ID'].values:
                        select_ids.append(select_id)
                        break
                    else:
                        print('Out of range. Try again.')
                except:
                    print('Invalid input. Try again.')
                

    # Update Performance summary excel file based on select_ids
    # "Selected" will be 1 for selected cells
    for ID in select_ids:
        df_performances.loc[df_performances['ID'] == ID, 'Selected'] = 1

    # Save the data
    with pd.ExcelWriter(performance_summary_path) as writer:
        df_performances.to_excel(writer, sheet_name='Performances', index=False)

    # Insert hyperlink for csv files
    insert_hyperlink(performance_summary_path,jv_folder_dir,performance_csv_path)
    # Design Excel
    Design_excel(performance_summary_path)
    # fill bkg color based on the cell id
    fill_excel_bkg(performance_summary_path,indication_number=8)

    print('DONE')

    return


